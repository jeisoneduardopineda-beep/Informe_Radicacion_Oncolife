from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path
import json

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table as PdfTable, TableStyle
from reportlab.lib.styles import getSampleStyleSheet


BASE_COLUMNS = [
    "Doc Pcte",
    "Fech Doc",
    "Convenio",
    "Regimen",
    "APB",
    "Saldo Apb",
    "Fecha Radicado",
    "Tipo Contrato",
    "MES",
]

MONTH_ORDER = [
    "enero",
    "febrero",
    "marzo",
    "abril",
    "mayo",
    "junio",
    "julio",
    "agosto",
    "septiembre",
    "octubre",
    "noviembre",
    "diciembre",
]

DATA_DIR = Path("dashboard_data")
LATEST_FILE = DATA_DIR / "BASE_RADICACION_ULTIMA_CARGA.xlsx"
LATEST_RIPS_FILE = DATA_DIR / "RIPS_ULTIMA_CARGA.json"


st.set_page_config(
    page_title="Dashboard Radicacion",
    layout="wide",
    initial_sidebar_state="expanded",
)


def css() -> None:
    st.markdown(
        """
        <style>
        :root {
          --bg:#071823; --panel:#102733; --panel2:#0c202b; --line:#244652;
          --ink:#edf7f8; --muted:#9ab0b8; --blue:#0bb7b9; --amber:#f7b718;
        }
        .stApp { background:#071823; color:var(--ink); }
        [data-testid="stSidebar"] { background:#0c202b; border-right:1px solid var(--line); }
        h1, h2, h3 { color:#edf7f8; letter-spacing:0; }
        .muted { color:var(--muted); font-size:13px; }
        .kpi {
          background:linear-gradient(135deg,#073b43,#0c202b);
          border:1px solid #244652;
          padding:16px;
          min-height:105px;
        }
        .kpi span { display:block; color:#fff; font-size:12px; font-weight:800; text-transform:uppercase; }
        .kpi strong { display:block; color:#fff; font-size:25px; margin-top:8px; }
        .kpi em { display:block; color:#ff3154; font-style:normal; font-weight:700; margin-top:6px; }
        .panel {
          background:#102733;
          border:1px solid #244652;
          padding:14px;
          min-height:230px;
        }
        .bar-row {
          display:grid;
          grid-template-columns:minmax(170px, 1.4fr) 2fr 140px;
          gap:10px;
          align-items:center;
          min-height:30px;
        }
        .bar-label { overflow:hidden; white-space:nowrap; text-overflow:ellipsis; font-size:13px; }
        .bar-track { height:17px; background:#324b52; border:1px solid #405e66; }
        .bar-fill { height:100%; background:linear-gradient(90deg,#f7b718,#0bb7b9); }
        .bar-value { text-align:right; font-variant-numeric:tabular-nums; font-size:13px; font-weight:800; }
        .rank-row {
          display:grid;
          grid-template-columns:34px 1fr auto;
          gap:10px;
          align-items:center;
          padding:11px 0;
          border-bottom:1px solid rgba(255,255,255,.08);
        }
        .rank-row b {
          width:28px; height:28px; border-radius:50%;
          display:grid; place-items:center;
          background:#f7b718; color:#061823;
        }
        .rank-row span { overflow:hidden; white-space:nowrap; text-overflow:ellipsis; font-weight:700; }
        .rank-row strong { font-size:13px; }
        .month-chart {
          height:245px;
          display:flex;
          align-items:end;
          gap:10px;
          border-bottom:1px solid #244652;
          padding:18px 8px 0;
        }
        .month-col { flex:1; display:flex; flex-direction:column; align-items:center; min-width:45px; }
        .month-value {
          color:#9ab0b8;
          font-size:11px;
          margin-bottom:6px;
          writing-mode:vertical-rl;
          transform:rotate(180deg);
          max-height:78px;
        }
        .month-bar { width:72%; min-width:24px; background:linear-gradient(180deg,#f7b718,#0bb7b9); }
        .month-label { margin-top:8px; color:#d5e4e7; font-weight:800; }
        .stDownloadButton button, .stButton button {
          background:#f7b718 !important;
          color:#071823 !important;
          border:0 !important;
          font-weight:800 !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def money(value: float) -> str:
    return "$" + f"{float(value or 0):,.0f}".replace(",", ".")


def number(value: float) -> str:
    return f"{float(value or 0):,.0f}".replace(",", ".")


def normalize_text(value: object) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def load_dataframe(file_or_path) -> pd.DataFrame:
    df = pd.read_excel(file_or_path, sheet_name="BASE")
    missing = [col for col in BASE_COLUMNS if col not in df.columns]
    if missing:
        raise ValueError("Faltan columnas en la hoja BASE: " + ", ".join(missing))

    df = df[BASE_COLUMNS].copy()
    df["Saldo Apb"] = pd.to_numeric(df["Saldo Apb"], errors="coerce").fillna(0)
    df["Fecha Radicado"] = pd.to_datetime(df["Fecha Radicado"], errors="coerce")
    df["Fech Doc"] = pd.to_datetime(df["Fech Doc"], errors="coerce", dayfirst=True)
    for col in ["Doc Pcte", "Convenio", "Regimen", "Tipo Contrato", "MES"]:
        df[col] = df[col].map(normalize_text)
    df["APB"] = df["APB"].map(normalize_text)
    df["MES"] = df["MES"].str.lower()
    return df


def service_date(service_type: str, service: dict) -> object:
    for key in ["fechaInicioAtencion", "fechaDispensAdmon", "fechaSuministroTecnologia", "fechaEgreso"]:
        if service.get(key):
            return service.get(key)
    return None


def service_code(service_type: str, service: dict) -> str:
    candidates = {
        "consultas": "codConsulta",
        "procedimientos": "codProcedimiento",
        "medicamentos": "codTecnologiaSalud",
        "otrosServicios": "codTecnologiaSalud",
    }
    return normalize_text(service.get(candidates.get(service_type, "")))


def service_name(service: dict) -> str:
    return normalize_text(service.get("nomTecnologiaSalud") or service.get("codTecnologiaSalud") or "")


def service_quantity(service_type: str, service: dict) -> float:
    for key in ["cantidadMedicamento", "cantidadOS"]:
        if key in service:
            try:
                return float(service.get(key) or 0)
            except Exception:
                return 0
    return 1


def load_rips_dataframe(file_or_path) -> tuple[pd.DataFrame, dict]:
    if hasattr(file_or_path, "read"):
        data = json.load(file_or_path)
    else:
        with open(file_or_path, encoding="utf-8-sig") as fh:
            data = json.load(fh)
    if not isinstance(data, dict) or "usuarios" not in data:
        raise ValueError("El JSON no parece tener estructura RIPS: falta la clave 'usuarios'.")

    rows = []
    for user in data.get("usuarios", []):
        user_doc = normalize_text(user.get("numDocumentoIdentificacion"))
        services = user.get("servicios") or {}
        for service_type, service_list in services.items():
            if not isinstance(service_list, list):
                continue
            for service in service_list:
                value = service.get("vrServicio")
                try:
                    value = float(value or 0)
                except Exception:
                    value = 0
                rows.append(
                    {
                        "Factura": normalize_text(data.get("numFactura")),
                        "Documento Usuario": user_doc,
                        "Tipo Documento": normalize_text(user.get("tipoDocumentoIdentificacion")),
                        "Tipo Usuario": normalize_text(user.get("tipoUsuario")),
                        "Sexo": normalize_text(user.get("codSexo")),
                        "Fecha Nacimiento": pd.to_datetime(user.get("fechaNacimiento"), errors="coerce"),
                        "Municipio": normalize_text(user.get("codMunicipioResidencia")),
                        "Zona": normalize_text(user.get("codZonaTerritorialResidencia")),
                        "Tipo Servicio": service_type,
                        "Fecha Servicio": pd.to_datetime(service_date(service_type, service), errors="coerce"),
                        "Prestador": normalize_text(service.get("codPrestador")),
                        "Codigo Servicio": service_code(service_type, service),
                        "Nombre Tecnologia": service_name(service),
                        "Diagnostico Principal": normalize_text(service.get("codDiagnosticoPrincipal")),
                        "Autorizacion": normalize_text(service.get("numAutorizacion")),
                        "Vr Servicio": value,
                        "Cantidad": service_quantity(service_type, service),
                        "Concepto Recaudo": normalize_text(service.get("conceptoRecaudo")),
                        "Consecutivo Servicio": service.get("consecutivo"),
                    }
                )
    df = pd.DataFrame(rows)
    if df.empty:
        raise ValueError("El JSON RIPS no contiene servicios para graficar.")
    metadata = {
        "numDocumentoIdObligado": normalize_text(data.get("numDocumentoIdObligado")),
        "numFactura": normalize_text(data.get("numFactura")),
        "usuarios": len(data.get("usuarios", [])),
    }
    return df, metadata


@st.cache_data(show_spinner=False)
def load_from_disk(path: str, mtime: float) -> pd.DataFrame:
    return load_dataframe(path)


@st.cache_data(show_spinner=False)
def load_rips_from_disk(path: str, mtime: float) -> tuple[pd.DataFrame, dict]:
    return load_rips_dataframe(path)


def latest_source() -> tuple[str, Path] | None:
    candidates = []
    if LATEST_FILE.exists():
        candidates.append(("radicacion", LATEST_FILE, LATEST_FILE.stat().st_mtime))
    if LATEST_RIPS_FILE.exists():
        candidates.append(("rips", LATEST_RIPS_FILE, LATEST_RIPS_FILE.stat().st_mtime))
    if not candidates:
        return None
    kind, path, _ = sorted(candidates, key=lambda item: item[2], reverse=True)[0]
    return kind, path


def get_data() -> tuple[str, pd.DataFrame, dict] | None:
    if "df" in st.session_state and "data_kind" in st.session_state:
        return st.session_state["data_kind"], st.session_state["df"], st.session_state.get("metadata", {})
    source = latest_source()
    if not source:
        return None
    kind, path = source
    if kind == "rips":
        df, metadata = load_rips_from_disk(str(path), path.stat().st_mtime)
        return kind, df, metadata
    return kind, load_from_disk(str(path), path.stat().st_mtime), {}


def save_uploaded(uploaded_file) -> pd.DataFrame:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    payload = uploaded_file.getvalue()
    filename = uploaded_file.name.lower()
    if filename.endswith(".json"):
        LATEST_RIPS_FILE.write_bytes(payload)
        df, metadata = load_rips_dataframe(BytesIO(payload))
        st.session_state["data_kind"] = "rips"
        st.session_state["metadata"] = metadata
        load_rips_from_disk.clear()
    else:
        LATEST_FILE.write_bytes(payload)
        df = load_dataframe(BytesIO(payload))
        st.session_state["data_kind"] = "radicacion"
        st.session_state["metadata"] = {}
        load_from_disk.clear()
    st.session_state["df"] = df
    return df


def group_sum(df: pd.DataFrame, field: str, top: int | None = None) -> list[dict]:
    grouped = df.groupby(field, dropna=False)["Saldo Apb"].sum().sort_values(ascending=False).reset_index()
    if top:
        grouped = grouped.head(top)
    return [{"label": str(row[field]) or "Sin dato", "value": float(row["Saldo Apb"])} for _, row in grouped.iterrows()]


def month_summary(df: pd.DataFrame) -> list[dict]:
    rows = []
    for month in MONTH_ORDER:
        piece = df[df["MES"] == month]
        if len(piece) or month in set(df["MES"].dropna()):
            rows.append({"label": month.capitalize(), "value": float(piece["Saldo Apb"].sum()), "count": int(piece["Doc Pcte"].count())})
    return rows


def apply_filters(df: pd.DataFrame, filters: dict[str, str]) -> pd.DataFrame:
    out = df
    if filters.get("mes"):
        out = out[out["MES"] == filters["mes"]]
    if filters.get("regimen"):
        out = out[out["Regimen"] == filters["regimen"]]
    if filters.get("contrato"):
        out = out[out["Tipo Contrato"] == filters["contrato"]]
    if filters.get("apb"):
        out = out[out["APB"] == filters["apb"]]
    if filters.get("convenio"):
        needle = filters["convenio"].lower()
        out = out[out["Convenio"].str.lower().str.contains(needle, na=False)]
    return out


def build_summary(df: pd.DataFrame) -> dict:
    total = float(df["Saldo Apb"].sum())
    docs = int(df["Doc Pcte"].count())
    convenios = int(df["Convenio"].replace("", pd.NA).dropna().nunique())
    apbs = int(df["APB"].replace("", pd.NA).dropna().nunique())
    return {
        "kpis": {
            "saldo_total": total,
            "documentos": docs,
            "convenios": convenios,
            "apb": apbs,
            "promedio_documento": total / docs if docs else 0,
        },
        "months": month_summary(df),
        "top_convenios": group_sum(df, "Convenio", 10),
        "top_apb": group_sum(df, "APB", 10),
        "regimen": group_sum(df, "Regimen"),
        "tipo_contrato": group_sum(df, "Tipo Contrato"),
    }


def rips_group_count(df: pd.DataFrame, field: str, top: int | None = None) -> list[dict]:
    grouped = df.groupby(field, dropna=False).size().sort_values(ascending=False).reset_index(name="Cantidad")
    if top:
        grouped = grouped.head(top)
    return [{"label": str(row[field]) or "Sin dato", "value": int(row["Cantidad"])} for _, row in grouped.iterrows()]


def rips_month_summary(df: pd.DataFrame) -> list[dict]:
    valid = df.dropna(subset=["Fecha Servicio"]).copy()
    if valid.empty:
        return []
    valid["Mes"] = valid["Fecha Servicio"].dt.to_period("M").astype(str)
    grouped = valid.groupby("Mes").size().reset_index(name="Cantidad")
    return [{"label": row["Mes"], "value": int(row["Cantidad"]), "count": int(row["Cantidad"])} for _, row in grouped.iterrows()]


def build_rips_summary(df: pd.DataFrame, metadata: dict) -> dict:
    total_services = int(len(df))
    users = int(df["Documento Usuario"].replace("", pd.NA).dropna().nunique())
    providers = int(df["Prestador"].replace("", pd.NA).dropna().nunique())
    dx = int(df["Diagnostico Principal"].replace("", pd.NA).dropna().nunique())
    total_value = float(df["Vr Servicio"].sum())
    return {
        "metadata": metadata,
        "kpis": {
            "usuarios": users,
            "servicios": total_services,
            "prestadores": providers,
            "diagnosticos": dx,
            "valor_total": total_value,
        },
        "months": rips_month_summary(df),
        "tipo_servicio": rips_group_count(df, "Tipo Servicio"),
        "prestadores": rips_group_count(df, "Prestador", 10),
        "diagnosticos": rips_group_count(df, "Diagnostico Principal", 10),
        "tecnologias": rips_group_count(df[df["Nombre Tecnologia"] != ""], "Nombre Tecnologia", 10),
        "sexo": rips_group_count(df, "Sexo"),
        "tipo_usuario": rips_group_count(df, "Tipo Usuario"),
    }


def bars(items: list[dict], limit: int = 10, value_kind: str = "money") -> str:
    rows = items[:limit]
    if not rows:
        return "<p class='muted'>Sin datos</p>"
    max_value = max(float(item["value"]) for item in rows) or 1
    html_rows = []
    for item in rows:
        width = max(4, min(100, float(item["value"]) / max_value * 100))
        display_value = money(item["value"]) if value_kind == "money" else number(item["value"])
        html_rows.append(
            f"<div class='bar-row'>"
            f"<div class='bar-label' title='{item['label']}'>{item['label']}</div>"
            f"<div class='bar-track'><div class='bar-fill' style='width:{width:.1f}%'></div></div>"
            f"<div class='bar-value'>{display_value}</div>"
            f"</div>"
        )
    return "".join(html_rows)


def month_chart(items: list[dict], value_kind: str = "money") -> str:
    if not items:
        return "<p class='muted'>Sin datos por mes</p>"
    max_value = max(float(item["value"]) for item in items) or 1
    cols = []
    for item in items:
        height = max(8, min(165, float(item["value"]) / max_value * 165))
        display_value = money(item["value"]) if value_kind == "money" else number(item["value"])
        cols.append(
            f"<div class='month-col'>"
            f"<div class='month-value'>{display_value}</div>"
            f"<div class='month-bar' style='height:{height:.0f}px'></div>"
            f"<div class='month-label'>{item['label'][:3]}</div>"
            f"</div>"
        )
    return "<div class='month-chart'>" + "".join(cols) + "</div>"


def kpi_card(label: str, value: str, foot: str = "") -> None:
    st.markdown(f"<div class='kpi'><span>{label}</span><strong>{value}</strong><em>{foot}</em></div>", unsafe_allow_html=True)


def make_excel_report(df: pd.DataFrame, summary: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Ejecutivo"
    ws.sheet_view.showGridLines = False
    dark = "071823"
    panel = "102733"
    teal = "0BB7B9"
    amber = "F7B718"
    soft = "EAF4F5"
    line = Side(style="thin", color="244652")

    for col in range(1, 14):
        ws.column_dimensions[get_column_letter(col)].width = 16
    for row in range(1, 42):
        ws.row_dimensions[row].height = 22

    ws.merge_cells("A1:M3")
    ws["A1"] = "Dashboard de Radicacion"
    ws["A1"].font = Font(size=28, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=dark)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A4:M4")
    ws["A4"] = f"Reporte ejecutivo generado el {datetime.now():%Y-%m-%d %H:%M}"
    ws["A4"].font = Font(size=11, color="D6E8EB", italic=True)
    ws["A4"].fill = PatternFill("solid", fgColor=dark)
    ws["A4"].alignment = Alignment(horizontal="center")

    k = summary["kpis"]
    kpi_cards = [
        ("A6:C9", "Saldo total", k["saldo_total"], "$#,##0", teal),
        ("D6:F9", "Documentos", k["documentos"], "#,##0", panel),
        ("G6:I9", "Convenios", k["convenios"], "#,##0", panel),
        ("J6:M9", "Promedio documento", k["promedio_documento"], "$#,##0", amber),
    ]
    for rng, label, value, fmt, color in kpi_cards:
        ws.merge_cells(rng)
        top_left = rng.split(":")[0]
        cell = ws[top_left]
        cell.value = value
        cell.number_format = fmt
        cell.font = Font(size=18, bold=True, color="071823" if color == amber else "FFFFFF")
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=line, right=line, top=line, bottom=line)
        label_cell = ws.cell(row=5, column=ws[top_left].column)
        label_cell.value = label.upper()
        label_cell.font = Font(size=10, bold=True, color="FFFFFF")
        label_cell.fill = PatternFill("solid", fgColor=dark)
        label_cell.alignment = Alignment(horizontal="center")

    def write_table(start_row: int, start_col: int, title: str, rows: list[dict], label_header: str = "Categoria") -> None:
        ws.cell(start_row, start_col, title)
        ws.cell(start_row, start_col).font = Font(size=13, bold=True, color="FFFFFF")
        ws.cell(start_row, start_col).fill = PatternFill("solid", fgColor=panel)
        ws.cell(start_row + 1, start_col, label_header)
        ws.cell(start_row + 1, start_col + 1, "Saldo")
        for c in range(start_col, start_col + 2):
            cell = ws.cell(start_row + 1, c)
            cell.font = Font(bold=True, color="071823")
            cell.fill = PatternFill("solid", fgColor=amber)
            cell.alignment = Alignment(horizontal="center")
        for idx, item in enumerate(rows, start=start_row + 2):
            ws.cell(idx, start_col, item["label"])
            ws.cell(idx, start_col + 1, item["value"])
            ws.cell(idx, start_col + 1).number_format = "$#,##0"
            for c in range(start_col, start_col + 2):
                cell = ws.cell(idx, c)
                cell.fill = PatternFill("solid", fgColor=soft if idx % 2 else "FFFFFF")
                cell.border = Border(bottom=Side(style="thin", color="D6E8EB"))
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    write_table(12, 1, "Saldo por mes", summary["months"], "Mes")
    write_table(12, 4, "Top convenios", summary["top_convenios"], "Convenio")
    write_table(12, 8, "Top APB", summary["top_apb"], "APB")

    line = LineChart()
    line.title = "Saldo por mes"
    line.add_data(Reference(ws, min_col=2, min_row=13, max_row=13 + len(summary["months"])), titles_from_data=True)
    line.set_categories(Reference(ws, min_col=1, min_row=14, max_row=13 + len(summary["months"])))
    line.height = 8
    line.width = 12
    ws.add_chart(line, "A26")

    bar = BarChart()
    bar.type = "bar"
    bar.title = "Top convenios"
    bar.add_data(Reference(ws, min_col=5, min_row=13, max_row=23), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=4, min_row=14, max_row=23))
    bar.height = 8
    bar.width = 15
    ws.add_chart(bar, "F26")

    base = wb.create_sheet("BASE")
    base.append(BASE_COLUMNS)
    for row in df.itertuples(index=False):
        base.append([x.to_pydatetime() if hasattr(x, "to_pydatetime") else x for x in row])
    ref = f"A1:{get_column_letter(len(BASE_COLUMNS))}{base.max_row}"
    tab = Table(displayName="Tabla_Base_Radicacion", ref=ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    base.add_table(tab)
    base.freeze_panes = "A2"
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center")

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def pdf_text(value: object) -> str:
    return ("" if value is None else str(value)).encode("latin-1", "replace").decode("latin-1")


def pdf_money(value: float) -> str:
    return "$" + f"{float(value or 0):,.0f}"


def pdf_table(title: str, rows: list[list], widths: list[float] | None = None) -> list:
    styles = getSampleStyleSheet()
    heading = styles["Heading3"]
    heading.textColor = colors.HexColor("#0B3B43")
    elements = [Paragraph(pdf_text(title), heading)]
    table = PdfTable([[pdf_text(cell) for cell in row] for row in rows], colWidths=widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0b3b43")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 8.5),
                ("FONTSIZE", (0, 1), (-1, -1), 7.5),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D6E8EB")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F7F8")]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (-1, 1), (-1, -1), "RIGHT"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    elements.extend([table, Spacer(1, 0.18 * inch)])
    return elements


def make_pdf_report(df: pd.DataFrame, summary: dict, filters: dict[str, str]) -> bytes:
    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(letter),
        leftMargin=0.35 * inch,
        rightMargin=0.35 * inch,
        topMargin=0.35 * inch,
        bottomMargin=0.35 * inch,
    )
    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    title_style.textColor = colors.white
    title_style.fontSize = 24
    subtitle_style = styles["Normal"]
    subtitle_style.textColor = colors.HexColor("#D6E8EB")
    subtitle_style.fontSize = 9
    filter_text = ", ".join(f"{k}: {v}" for k, v in filters.items() if v) or "Todos"

    story = []
    hero = PdfTable(
        [
            [Paragraph("Dashboard de Radicacion", title_style)],
            [Paragraph(pdf_text(f"Reporte ejecutivo | Generado: {datetime.now():%Y-%m-%d %H:%M} | Filtros: {filter_text}"), subtitle_style)],
        ],
        colWidths=[10.2 * inch],
        rowHeights=[0.55 * inch, 0.35 * inch],
    )
    hero.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#071823")),
                ("LEFTPADDING", (0, 0), (-1, -1), 16),
                ("RIGHTPADDING", (0, 0), (-1, -1), 16),
                ("TOPPADDING", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    story.extend([hero, Spacer(1, 0.16 * inch)])

    k = summary["kpis"]
    kpi_data = [
        ["SALDO TOTAL", "DOCUMENTOS", "CONVENIOS", "PROMEDIO DOC."],
        [pdf_money(k["saldo_total"]), number(k["documentos"]), number(k["convenios"]), pdf_money(k["promedio_documento"])],
    ]
    kpi_table = PdfTable(kpi_data, colWidths=[2.55 * inch] * 4, rowHeights=[0.28 * inch, 0.46 * inch])
    kpi_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0B3B43")),
                ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#102733")),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 8),
                ("FONTSIZE", (0, 1), (-1, 1), 14),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.6, colors.HexColor("#244652")),
            ]
        )
    )
    story.extend([kpi_table, Spacer(1, 0.18 * inch)])
    story.extend(pdf_table("Saldo por mes", [["Mes", "Saldo", "Documentos"]] + [[i["label"], pdf_money(i["value"]), number(i["count"])] for i in summary["months"]], [1.5 * inch, 1.8 * inch, 1.2 * inch]))
    story.extend(pdf_table("Top convenios", [["Convenio", "Saldo"]] + [[i["label"][:70], pdf_money(i["value"])] for i in summary["top_convenios"]], [5.0 * inch, 1.6 * inch]))
    story.extend(pdf_table("Top APB", [["APB", "Saldo"]] + [[i["label"], pdf_money(i["value"])] for i in summary["top_apb"]], [1.6 * inch, 1.8 * inch]))
    detail = df.sort_values("Saldo Apb", ascending=False).head(35)
    rows = [["Documento", "Fecha", "Convenio", "Regimen", "APB", "Contrato", "Mes", "Saldo"]]
    for _, row in detail.iterrows():
        rows.append(
            [
                row["Doc Pcte"],
                row["Fecha Radicado"].strftime("%Y-%m-%d") if not pd.isna(row["Fecha Radicado"]) else "",
                row["Convenio"][:38],
                row["Regimen"][:32],
                row["APB"],
                row["Tipo Contrato"],
                str(row["MES"]).capitalize(),
                pdf_money(row["Saldo Apb"]),
            ]
        )
    story.extend(pdf_table("Detalle principal", rows, [0.85 * inch, 0.75 * inch, 2.2 * inch, 1.8 * inch, 0.65 * inch, 0.75 * inch, 0.55 * inch, 1.1 * inch]))
    doc.build(story)
    return output.getvalue()


def make_rips_excel_report(df: pd.DataFrame, summary: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte RIPS"
    ws.sheet_view.showGridLines = False
    dark = "071823"
    panel = "102733"
    teal = "0BB7B9"
    amber = "F7B718"
    soft = "EAF4F5"
    line = Side(style="thin", color="244652")
    for col in range(1, 14):
        ws.column_dimensions[get_column_letter(col)].width = 16

    ws.merge_cells("A1:M3")
    ws["A1"] = "Dashboard RIPS"
    ws["A1"].font = Font(size=28, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=dark)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A4:M4")
    meta = summary.get("metadata", {})
    ws["A4"] = f"Factura {meta.get('numFactura','')} | Generado el {datetime.now():%Y-%m-%d %H:%M}"
    ws["A4"].font = Font(size=11, color="D6E8EB", italic=True)
    ws["A4"].fill = PatternFill("solid", fgColor=dark)
    ws["A4"].alignment = Alignment(horizontal="center")

    k = summary["kpis"]
    cards = [
        ("A6:C9", "Usuarios", k["usuarios"], "#,##0", teal),
        ("D6:F9", "Servicios", k["servicios"], "#,##0", panel),
        ("G6:I9", "Prestadores", k["prestadores"], "#,##0", panel),
        ("J6:M9", "Diagnosticos", k["diagnosticos"], "#,##0", amber),
    ]
    for rng, label, value, fmt, color in cards:
        ws.merge_cells(rng)
        cell = ws[rng.split(":")[0]]
        cell.value = value
        cell.number_format = fmt
        cell.font = Font(size=18, bold=True, color="071823" if color == amber else "FFFFFF")
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=line, right=line, top=line, bottom=line)
        label_cell = ws.cell(row=5, column=cell.column)
        label_cell.value = label.upper()
        label_cell.font = Font(size=10, bold=True, color="FFFFFF")
        label_cell.fill = PatternFill("solid", fgColor=dark)
        label_cell.alignment = Alignment(horizontal="center")

    def write_count_table(start_row: int, start_col: int, title: str, rows: list[dict], header: str) -> None:
        ws.cell(start_row, start_col, title)
        ws.cell(start_row, start_col).font = Font(size=13, bold=True, color="FFFFFF")
        ws.cell(start_row, start_col).fill = PatternFill("solid", fgColor=panel)
        ws.cell(start_row + 1, start_col, header)
        ws.cell(start_row + 1, start_col + 1, "Cantidad")
        for c in range(start_col, start_col + 2):
            h = ws.cell(start_row + 1, c)
            h.font = Font(bold=True, color="071823")
            h.fill = PatternFill("solid", fgColor=amber)
            h.alignment = Alignment(horizontal="center")
        for idx, item in enumerate(rows, start=start_row + 2):
            ws.cell(idx, start_col, item["label"])
            ws.cell(idx, start_col + 1, item["value"])
            ws.cell(idx, start_col + 1).number_format = "#,##0"
            for c in range(start_col, start_col + 2):
                cell = ws.cell(idx, c)
                cell.fill = PatternFill("solid", fgColor=soft if idx % 2 else "FFFFFF")
                cell.border = Border(bottom=Side(style="thin", color="D6E8EB"))
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    write_count_table(12, 1, "Servicios por tipo", summary["tipo_servicio"], "Tipo")
    write_count_table(12, 4, "Top diagnosticos", summary["diagnosticos"], "Diagnostico")
    write_count_table(12, 8, "Top tecnologias", summary["tecnologias"], "Tecnologia")

    detail = wb.create_sheet("Detalle Servicios")
    cols = [
        "Factura",
        "Documento Usuario",
        "Tipo Usuario",
        "Sexo",
        "Tipo Servicio",
        "Fecha Servicio",
        "Prestador",
        "Codigo Servicio",
        "Nombre Tecnologia",
        "Diagnostico Principal",
        "Vr Servicio",
        "Cantidad",
    ]
    detail.append(cols)
    for row in df[cols].itertuples(index=False):
        detail.append([x.to_pydatetime() if hasattr(x, "to_pydatetime") else x for x in row])
    ref = f"A1:{get_column_letter(len(cols))}{detail.max_row}"
    tab = Table(displayName="Tabla_RIPS_Servicios", ref=ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    detail.add_table(tab)
    detail.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def make_rips_pdf_report(df: pd.DataFrame, summary: dict, filters: dict[str, str]) -> bytes:
    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(letter),
        leftMargin=0.35 * inch,
        rightMargin=0.35 * inch,
        topMargin=0.35 * inch,
        bottomMargin=0.35 * inch,
    )
    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    title_style.textColor = colors.white
    title_style.fontSize = 24
    subtitle_style = styles["Normal"]
    subtitle_style.textColor = colors.HexColor("#D6E8EB")
    subtitle_style.fontSize = 9
    meta = summary.get("metadata", {})
    filter_text = ", ".join(f"{k}: {v}" for k, v in filters.items() if v) or "Todos"
    hero = PdfTable(
        [
            [Paragraph("Dashboard RIPS", title_style)],
            [Paragraph(pdf_text(f"Factura: {meta.get('numFactura','')} | Generado: {datetime.now():%Y-%m-%d %H:%M} | Filtros: {filter_text}"), subtitle_style)],
        ],
        colWidths=[10.2 * inch],
        rowHeights=[0.55 * inch, 0.35 * inch],
    )
    hero.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#071823")), ("LEFTPADDING", (0, 0), (-1, -1), 16)]))
    story = [hero, Spacer(1, 0.16 * inch)]
    k = summary["kpis"]
    kpi_table = PdfTable(
        [["USUARIOS", "SERVICIOS", "PRESTADORES", "DIAGNOSTICOS"], [number(k["usuarios"]), number(k["servicios"]), number(k["prestadores"]), number(k["diagnosticos"])]],
        colWidths=[2.55 * inch] * 4,
        rowHeights=[0.28 * inch, 0.46 * inch],
    )
    kpi_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0B3B43")),
                ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#102733")),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 8),
                ("FONTSIZE", (0, 1), (-1, 1), 14),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.6, colors.HexColor("#244652")),
            ]
        )
    )
    story.extend([kpi_table, Spacer(1, 0.18 * inch)])
    story.extend(pdf_table("Servicios por tipo", [["Tipo", "Cantidad"]] + [[i["label"], number(i["value"])] for i in summary["tipo_servicio"]], [2.2 * inch, 1.2 * inch]))
    story.extend(pdf_table("Top diagnosticos", [["Diagnostico", "Cantidad"]] + [[i["label"], number(i["value"])] for i in summary["diagnosticos"]], [2.2 * inch, 1.2 * inch]))
    story.extend(pdf_table("Top tecnologias", [["Tecnologia", "Cantidad"]] + [[i["label"][:70], number(i["value"])] for i in summary["tecnologias"]], [5.0 * inch, 1.2 * inch]))
    detail = df.sort_values("Fecha Servicio", ascending=False).head(35)
    rows = [["Usuario", "Fecha", "Tipo", "Codigo", "Tecnologia", "DX", "Prestador"]]
    for _, row in detail.iterrows():
        rows.append(
            [
                row["Documento Usuario"],
                row["Fecha Servicio"].strftime("%Y-%m-%d") if not pd.isna(row["Fecha Servicio"]) else "",
                row["Tipo Servicio"],
                row["Codigo Servicio"],
                row["Nombre Tecnologia"][:35],
                row["Diagnostico Principal"],
                row["Prestador"],
            ]
        )
    story.extend(pdf_table("Detalle principal", rows, [0.95 * inch, 0.8 * inch, 1.15 * inch, 0.8 * inch, 2.2 * inch, 0.65 * inch, 1.2 * inch]))
    doc.build(story)
    return output.getvalue()


def admin_panel() -> None:
    st.sidebar.markdown("### Administracion")
    admin_password = st.secrets.get("ADMIN_PASSWORD", "tu-clave-segura")
    typed = st.sidebar.text_input("Clave para cargar datos", type="password")
    if not typed:
        st.sidebar.caption("La vista publica puede filtrar y descargar, pero no cargar datos.")
        return
    if typed != admin_password:
        st.sidebar.error("Clave incorrecta. Revisa el valor ADMIN_PASSWORD en los secrets de Streamlit.")
        st.sidebar.caption("La vista publica puede filtrar y descargar, pero no cargar datos.")
        return
    st.sidebar.success("Clave correcta. Ya puedes cargar datos.")
    uploaded = st.sidebar.file_uploader("Cargar BASE_RADICACION o JSON RIPS", type=["xlsx", "xls", "json"])
    if uploaded:
        try:
            df = save_uploaded(uploaded)
            st.sidebar.success(f"Base cargada: {len(df):,} registros")
            st.rerun()
        except Exception as exc:
            st.sidebar.error(str(exc))


def render_dashboard(df_raw: pd.DataFrame) -> None:
    st.sidebar.markdown("### Filtros")
    months_present = set(df_raw["MES"].dropna().tolist())
    month_options = ["Todos"] + [m.capitalize() for m in MONTH_ORDER if m in months_present]
    regimen_options = ["Todos"] + sorted([x for x in df_raw["Regimen"].dropna().unique().tolist() if x])
    contrato_options = ["Todos"] + sorted([x for x in df_raw["Tipo Contrato"].dropna().unique().tolist() if x])
    apb_options = ["Todos"] + sorted([x for x in df_raw["APB"].dropna().unique().tolist() if x])

    mes_label = st.sidebar.selectbox("Mes", month_options)
    regimen = st.sidebar.selectbox("Regimen", regimen_options)
    contrato = st.sidebar.selectbox("Tipo contrato", contrato_options)
    apb = st.sidebar.selectbox("APB", apb_options)
    convenio = st.sidebar.text_input("Buscar convenio")

    filters = {
        "mes": "" if mes_label == "Todos" else mes_label.lower(),
        "regimen": "" if regimen == "Todos" else regimen,
        "contrato": "" if contrato == "Todos" else contrato,
        "apb": "" if apb == "Todos" else apb,
        "convenio": convenio.strip(),
    }
    df = apply_filters(df_raw, filters)
    summary = build_summary(df)
    k = summary["kpis"]

    st.title("Dashboard de Radicacion")
    st.markdown("<p class='muted'>Vista publica: las personas pueden filtrar, navegar y descargar reportes. Solo quien tenga clave puede cargar datos.</p>", unsafe_allow_html=True)

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        kpi_card("Saldo total", money(k["saldo_total"]), "con filtros activos")
    with c2:
        kpi_card("Documentos", number(k["documentos"]), "registros filtrados")
    with c3:
        kpi_card("Convenios", number(k["convenios"]), "entidades distintas")
    with c4:
        kpi_card("Promedio doc.", money(k["promedio_documento"]), "saldo / documentos")

    excel_bytes = make_excel_report(df, summary)
    pdf_bytes = make_pdf_report(df, summary, filters)
    d1, d2 = st.columns([1, 1])
    with d1:
        st.download_button("Descargar PDF", pdf_bytes, "reporte_radicacion.pdf", "application/pdf")
    with d2:
        st.download_button("Descargar Excel", excel_bytes, "reporte_radicacion.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    tab1, tab2, tab3, tab4, tab5 = st.tabs(["Resumen", "Convenios", "APB", "Regimen y contrato", "Base"])
    with tab1:
        left, right = st.columns([2, 1])
        with left:
            st.markdown("<div class='panel'><h3>Saldo por mes</h3>" + month_chart(summary["months"]) + "</div>", unsafe_allow_html=True)
            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown("<div class='panel'><h3>Top convenios</h3>" + bars(summary["top_convenios"], 10) + "</div>", unsafe_allow_html=True)
        with right:
            rank = []
            for idx, item in enumerate(summary["top_convenios"][:5], start=1):
                rank.append(
                    f"<div class='rank-row'><b>{idx}</b>"
                    f"<span>{item['label']}</span><strong>{money(item['value'])}</strong></div>"
                )
            st.markdown("<div class='panel'><h3>Top 5 entidades</h3>" + "".join(rank) + "</div>", unsafe_allow_html=True)
            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown("<div class='panel'><h3>Tipo de contrato</h3>" + bars(summary["tipo_contrato"], 10) + "</div>", unsafe_allow_html=True)
    with tab2:
        st.markdown("<div class='panel'><h3>Convenios por saldo</h3>" + bars(summary["top_convenios"], 10) + "</div>", unsafe_allow_html=True)
    with tab3:
        st.markdown("<div class='panel'><h3>APB por saldo</h3>" + bars(summary["top_apb"], 10) + "</div>", unsafe_allow_html=True)
    with tab4:
        left, right = st.columns(2)
        with left:
            st.markdown("<div class='panel'><h3>Regimen</h3>" + bars(summary["regimen"], 20) + "</div>", unsafe_allow_html=True)
        with right:
            st.markdown("<div class='panel'><h3>Tipo de contrato</h3>" + bars(summary["tipo_contrato"], 20) + "</div>", unsafe_allow_html=True)
    with tab5:
        view = df.sort_values("Saldo Apb", ascending=False).head(500).copy()
        view["Saldo Apb"] = view["Saldo Apb"].map(money)
        st.dataframe(view, use_container_width=True, height=520)


def render_rips_dashboard(df_raw: pd.DataFrame, metadata: dict) -> None:
    st.sidebar.markdown("### Filtros RIPS")
    tipo_options = ["Todos"] + sorted([x for x in df_raw["Tipo Servicio"].dropna().unique().tolist() if x])
    prestador_options = ["Todos"] + sorted([x for x in df_raw["Prestador"].dropna().unique().tolist() if x])
    sexo_options = ["Todos"] + sorted([x for x in df_raw["Sexo"].dropna().unique().tolist() if x])
    dx_options = ["Todos"] + sorted([x for x in df_raw["Diagnostico Principal"].dropna().unique().tolist() if x])

    tipo_servicio = st.sidebar.selectbox("Tipo servicio", tipo_options)
    prestador = st.sidebar.selectbox("Prestador", prestador_options)
    sexo = st.sidebar.selectbox("Sexo", sexo_options)
    diagnostico = st.sidebar.selectbox("Diagnostico", dx_options)
    busqueda = st.sidebar.text_input("Buscar tecnologia o usuario")

    df = df_raw
    filters = {
        "tipo_servicio": "" if tipo_servicio == "Todos" else tipo_servicio,
        "prestador": "" if prestador == "Todos" else prestador,
        "sexo": "" if sexo == "Todos" else sexo,
        "diagnostico": "" if diagnostico == "Todos" else diagnostico,
        "busqueda": busqueda.strip(),
    }
    if filters["tipo_servicio"]:
        df = df[df["Tipo Servicio"] == filters["tipo_servicio"]]
    if filters["prestador"]:
        df = df[df["Prestador"] == filters["prestador"]]
    if filters["sexo"]:
        df = df[df["Sexo"] == filters["sexo"]]
    if filters["diagnostico"]:
        df = df[df["Diagnostico Principal"] == filters["diagnostico"]]
    if filters["busqueda"]:
        needle = filters["busqueda"].lower()
        df = df[
            df["Nombre Tecnologia"].str.lower().str.contains(needle, na=False)
            | df["Documento Usuario"].str.lower().str.contains(needle, na=False)
            | df["Codigo Servicio"].str.lower().str.contains(needle, na=False)
        ]

    summary = build_rips_summary(df, metadata)
    k = summary["kpis"]
    st.title("Dashboard RIPS")
    st.markdown(
        f"<p class='muted'>Factura {metadata.get('numFactura','')} | Obligado {metadata.get('numDocumentoIdObligado','')} | Datos de servicios RIPS normalizados.</p>",
        unsafe_allow_html=True,
    )

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        kpi_card("Usuarios", number(k["usuarios"]), "unicos")
    with c2:
        kpi_card("Servicios", number(k["servicios"]), "registros")
    with c3:
        kpi_card("Prestadores", number(k["prestadores"]), "codigos")
    with c4:
        kpi_card("Diagnosticos", number(k["diagnosticos"]), "principales")

    pdf_bytes = make_rips_pdf_report(df, summary, filters)
    excel_bytes = make_rips_excel_report(df, summary)
    d1, d2 = st.columns([1, 1])
    with d1:
        st.download_button("Descargar PDF RIPS", pdf_bytes, "reporte_rips.pdf", "application/pdf")
    with d2:
        st.download_button("Descargar Excel RIPS", excel_bytes, "reporte_rips.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    tab1, tab2, tab3, tab4, tab5 = st.tabs(["Resumen", "Servicios", "Diagnosticos", "Tecnologias", "Detalle"])
    with tab1:
        left, right = st.columns([2, 1])
        with left:
            st.markdown("<div class='panel'><h3>Servicios por fecha</h3>" + month_chart(summary["months"], "count") + "</div>", unsafe_allow_html=True)
            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown("<div class='panel'><h3>Servicios por tipo</h3>" + bars(summary["tipo_servicio"], 10, "count") + "</div>", unsafe_allow_html=True)
        with right:
            st.markdown("<div class='panel'><h3>Sexo</h3>" + bars(summary["sexo"], 10, "count") + "</div>", unsafe_allow_html=True)
            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown("<div class='panel'><h3>Tipo usuario</h3>" + bars(summary["tipo_usuario"], 10, "count") + "</div>", unsafe_allow_html=True)
    with tab2:
        st.markdown("<div class='panel'><h3>Prestadores</h3>" + bars(summary["prestadores"], 10, "count") + "</div>", unsafe_allow_html=True)
    with tab3:
        st.markdown("<div class='panel'><h3>Diagnosticos principales</h3>" + bars(summary["diagnosticos"], 10, "count") + "</div>", unsafe_allow_html=True)
    with tab4:
        st.markdown("<div class='panel'><h3>Tecnologias / servicios</h3>" + bars(summary["tecnologias"], 10, "count") + "</div>", unsafe_allow_html=True)
    with tab5:
        cols = [
            "Factura",
            "Documento Usuario",
            "Tipo Usuario",
            "Sexo",
            "Tipo Servicio",
            "Fecha Servicio",
            "Prestador",
            "Codigo Servicio",
            "Nombre Tecnologia",
            "Diagnostico Principal",
            "Vr Servicio",
            "Cantidad",
        ]
        st.dataframe(df[cols].sort_values("Fecha Servicio", ascending=False).head(1000), use_container_width=True, height=520)


def main() -> None:
    css()
    admin_panel()
    loaded = get_data()
    if loaded is None:
        st.title("Dashboard de Radicacion")
        st.info("Aun no hay una base cargada. Ingresa la clave de administracion en el panel lateral y carga un Excel de radicacion o un JSON RIPS.")
        return
    kind, df, metadata = loaded
    if kind == "rips":
        render_rips_dashboard(df, metadata)
    else:
        render_dashboard(df)


if __name__ == "__main__":
    main()
